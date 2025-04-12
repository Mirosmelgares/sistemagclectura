from django.shortcuts import render, redirect
from django.db.models import Count, Sum, Q
from django.http import JsonResponse
from django.template.loader import render_to_string
from ..models import RegistroUsuarios, EvaluacionLecturaIndividual

VAL_MAX = 3

# Vista: Dashboard del administrador
def dashboard_admin(request):
    usuarios_activos = RegistroUsuarios.objects.annotate(
        evaluaciones_realizadas=Count(
            'evaluacionlecturaindividual',
            filter=Q(evaluacionlecturaindividual__puntaje__isnull=False)
        )
    ).filter(evaluaciones_realizadas__gt=0)

    total_usuarios = usuarios_activos.count()

    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]
    resumen = []

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(tipo_texto=tipo)
        total_lecturas = lecturas.count()
        total_puntos = sum([e.puntaje for e in lecturas if e.puntaje is not None])
        promedio = (total_puntos / total_lecturas) if total_lecturas > 0 else 0
        porcentaje = (promedio / VAL_MAX) * 100 if total_lecturas > 0 else 0

        resumen.append({
            "tipo": tipo,
            "total_lecturas": total_lecturas,
            "promedio_porcentaje": round(porcentaje, 2)
        })

    return render(request, "evaluacionescl/dashboard_admin.html", {
        "total_usuarios": total_usuarios,
        "resumen": resumen
    })

# Vista: Resultados globales por tipo de texto
def admin_resultados(request):
    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]
    resumen = []

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(tipo_texto=tipo)
        cantidad = lecturas.count()
        suma_puntos = sum([e.puntaje for e in lecturas if e.puntaje is not None])

        if cantidad > 0:
            promedio = suma_puntos / cantidad
            porcentaje = (promedio / VAL_MAX) * 100  # Asumiendo VAL_MAX = 3

            if porcentaje >= 89:
                nivel = f"Alto (Comprensión profunda) - {int(porcentaje)}%"
            elif porcentaje >= 67:
                nivel = f"Medio (Comprensión adecuada) - {int(porcentaje)}%"
            elif porcentaje >= 34:
                nivel = f"Bajo (Comprensión superficial) - {int(porcentaje)}%"
            else:
                nivel = f"Deficiente (No comprensión) - {int(porcentaje)}%"
        else:
            nivel = "Sin evaluar"

        resumen.append({
            "tipo": tipo,
            "cantidad": cantidad,
            "puntaje_total": suma_puntos,
            "nivel": nivel
        })

    total_usuarios = RegistroUsuarios.objects.count()

    return render(request, 'evaluacionescl/admin_resultados.html', {
        'resumen': resumen,
        'total_usuarios': total_usuarios
    })

# Vista: Estadísticas por alumno con búsqueda
from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from ..models import RegistroUsuarios, EvaluacionLecturaIndividual
from .evaluacion_views import calcular_porcentaje

VAL_MAX = 3

def admin_estadisticas(request):
    query = request.GET.get('q', '')
    usuarios = RegistroUsuarios.objects.all()
    if query:
        usuarios = usuarios.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(matricula__icontains=query)
        )

    resultados = []
    for user in usuarios:
        lecturas = EvaluacionLecturaIndividual.objects.filter(usuario=user)
        total_textos = lecturas.count()
        total_puntos = sum([e.puntaje for e in lecturas if e.puntaje is not None])
        promedio = (total_puntos / total_textos) if total_textos > 0 else 0
        porcentaje = (promedio / VAL_MAX) * 100 if total_textos > 0 else 0

        if total_textos > 0:
            if porcentaje >= 89:
                nivel = f"Alto (Comprensión profunda) - {int(porcentaje)}%"
            elif porcentaje >= 67:
                nivel = f"Medio (Comprensión adecuada) - {int(porcentaje)}%"
            elif porcentaje >= 34:
                nivel = f"Bajo (Comprensión superficial) - {int(porcentaje)}%"
            else:
                nivel = f"Deficiente (No comprensión) - {int(porcentaje)}%"
        else:
            nivel = "Sin evaluar"

        resultados.append({
            "id": user.id,
            "nombre": f"{user.nombre} {user.apellido}",
            "matricula": user.matricula,
            "textos": total_textos,
            "puntaje": total_puntos,
            "nivel": nivel
        })

    return render(request, 'evaluacionescl/admin_estadisticas.html', {
        'resultados': resultados,
        'query': query
    })
#--------------------------------------------------------------------------------------------

from django.contrib import messages
from django.shortcuts import redirect

def eliminar_usuario(request, usuario_id):
    if request.method == "POST":
        try:
            usuario = RegistroUsuarios.objects.get(id=usuario_id)
            usuario.delete()
            messages.success(request, "Usuario eliminado correctamente.")
        except RegistroUsuarios.DoesNotExist:
            messages.error(request, "El usuario no existe.")
    return redirect("admin_estadisticas")

#----------------------------------------------------------------------------------------------
from django.shortcuts import get_object_or_404, render, redirect
from ..models import RegistroUsuarios, EvaluacionLecturaIndividual
from .evaluacion_views import calcular_porcentaje  # Asegúrate de importar si está separado

def ver_resultados_alumno(request, usuario_id):
    alumno = get_object_or_404(RegistroUsuarios, id=usuario_id)
    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]
    resultados = []

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(usuario=alumno, tipo_texto=tipo)
        total = lecturas.count()

        if total:
            puntos = sum([e.puntaje for e in lecturas if e.puntaje is not None])
            promedio = puntos / total
            porcentaje = calcular_porcentaje(promedio)
            if porcentaje >= 89:
                nivel = f"Alto (Comprensión profunda) - {int(porcentaje)}%"
            elif porcentaje >= 67:
                nivel = f"Medio (Comprensión adecuada) - {int(porcentaje)}%"
            elif porcentaje >= 34:
                nivel = f"Bajo (Comprensión superficial) - {int(porcentaje)}%"
            else:
                nivel = f"Deficiente (No comprensión) - {int(porcentaje)}%"
        else:
            nivel = "Sin evaluar"

        resultados.append({
            "tipo": tipo,
            "cantidad": total,
            "nivel": nivel
        })

            # Si no tiene evaluaciones en ningún tipo, mostramos mensaje especial
    evaluaciones_total = sum([r["cantidad"] for r in resultados])
    if evaluaciones_total == 0:
        return render(request, "evaluacionescl/sin_resultados_alumno.html", {
            "mensaje": f"El alumno {alumno.nombre} aún no ha realizado ninguna evaluación."
        })


    return render(request, "evaluacionescl/ver_resultados_alumno.html", {
    "resultados": resultados,
    "alumno": alumno,
    "usuario_id": alumno.id,  # Agregado esto
    "nombre_alumno": alumno.nombre,
})



from django.shortcuts import render, get_object_or_404
from ..models import RegistroUsuarios, EvaluacionLecturaIndividual

def calcular_porcentaje(promedio):
    if promedio is None:
        return 0
    return (promedio / VAL_MAX) * 100

def ver_grafica_alumno_tipo(request, usuario_id, tipo_texto):
    alumno = get_object_or_404(RegistroUsuarios, id=usuario_id)
    lecturas = EvaluacionLecturaIndividual.objects.filter(
        usuario=alumno,
        tipo_texto=tipo_texto
    ).order_by('titulo_lectura', '-fecha_lectura')

    if not lecturas.exists():
        return render(request, "evaluacionescl/no_evaluaciones_admin.html", {
            "mensaje": f"No hay lecturas evaluadas para el tipo '{tipo_texto}' de este alumno."
        })

    titulos = []
    porcentajes = []
    tooltips = []
    tipos_inferencia = []
    vistos = set()

    for l in lecturas:
        if l.titulo_lectura not in vistos:
            vistos.add(l.titulo_lectura)
            titulos.append(l.titulo_lectura)

        porcentaje = calcular_porcentaje(l.puntaje or 0)
        tooltip = f"{int(porcentaje)}%" if l.puntaje is not None else "Sin evaluar"

        tipo = l.tipo_inferencia

        # Limpieza y asignación segura
        if tipo == "no_inferencia_sinsentido":
                tipo = "No inferencia: sin sentido"
        elif tipo == "no_inferencia_parafrasis":
                tipo = "No inferencia: paráfrasis"
        elif tipo:
                tipo = tipo.capitalize()
        else:
                tipo = "No inferencia: sin sentido"  # Por defecto si está vacío o None

        porcentajes.append(porcentaje)
        tooltips.append(tooltip)
        tipos_inferencia.append(tipo)

    return render(request, "evaluacionescl/ver_grafica_alumno_tipo.html", {
        "tipo": tipo_texto,
        "titulos": titulos,
        "porcentajes": porcentajes,
        "tooltips": tooltips,
        "tipos_inferencia": tipos_inferencia, 
        "alumno": alumno,
        "alumno_id": alumno.id #clave
    })


# Exportar estadisticas admin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from ..models import EvaluacionLecturaIndividual

def exportar_admin_estadisticas_excel(request):
    from ..models import RegistroUsuarios
    from .evaluacion_views import calcular_porcentaje

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estadísticas por alumno"

    headers = ["Nombre", "Matrícula", "Textos leídos", "Puntaje total", "Porcentaje", "Nivel de comprensión"]
    ws.append(headers)

    usuarios = RegistroUsuarios.objects.all()
    for user in usuarios:
        lecturas = EvaluacionLecturaIndividual.objects.filter(usuario=user)
        total = lecturas.count()
        puntaje_total = sum([e.puntaje for e in lecturas if e.puntaje is not None])
        porcentaje = calcular_porcentaje(puntaje_total / total) if total else 0

        if total:
            if porcentaje >= 89:
                nivel = "Alto (Comprensión profunda)"
            elif porcentaje >= 67:
                nivel = "Medio (Comprensión adecuada)"
            elif porcentaje >= 34:
                nivel = "Bajo (Comprensión superficial)"
            else:
                nivel = "Deficiente (No comprensión)"
        else:
            nivel = "Sin evaluar"

        fila = [
            f"{user.nombre} {user.apellido}",
            user.matricula,
            total,
            puntaje_total,
            int(porcentaje),
            nivel
        ]
        ws.append(fila)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=admin_estadisticas_alumnos.xlsx'
    wb.save(response)
    return response


# Vista para exportar resultados globales
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from ..models import EvaluacionLecturaIndividual

VAL_MAX = 3  # Si no está definido arriba, agrégalo

def exportar_admin_resultados_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen por tipo de texto"

    headers = ["Tipo de texto", "Documentos leídos", "Puntaje total", "Porcentaje", "Nivel de comprensión"]
    ws.append(headers)

    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(tipo_texto=tipo)
        cantidad = lecturas.count()
        suma = sum([e.puntaje for e in lecturas if e.puntaje is not None])
        promedio = (suma / cantidad) if cantidad > 0 else 0
        porcentaje = (promedio / VAL_MAX * 100) if cantidad > 0 else 0

        if cantidad > 0:
            if porcentaje >= 89:
                nivel = "Alto (Comprensión profunda)"
            elif porcentaje >= 67:
                nivel = "Medio (Comprensión adecuada)"
            elif porcentaje >= 34:
                nivel = "Bajo (Comprensión superficial)"
            else:
                nivel = "Deficiente (No comprensión)"
        else:
            nivel = "Sin evaluar"

        fila = [tipo, cantidad, suma, int(porcentaje), nivel]
        ws.append(fila)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=admin_resultados_globales.xlsx'
    wb.save(response)
    return response



#-----------------------------------------------------------------------------
from django.db import connection
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.messages import add_message
from ..models import RegistroAdmin, RegistroUsuarios, EvaluacionLectura, EvaluacionLecturaIndividual, VistaAdmin

def reset_auto_increment(tabla):
    with connection.cursor() as cursor:
        cursor.execute(f"ALTER TABLE {tabla} AUTO_INCREMENT = 1")

def resetear_datos(request):
    if 'admin_id' not in request.session:
        return redirect('login_admin')

    if request.method == "POST":
        if "reset_todo" in request.POST:
            # Eliminar todo, incluyendo el admin actual
            RegistroUsuarios.objects.all().delete()
            RegistroAdmin.objects.all().delete()
            EvaluacionLectura.objects.all().delete()
            EvaluacionLecturaIndividual.objects.all().delete()
            VistaAdmin.objects.all().delete()

            reset_auto_increment("evaluacionescl_registrousuarios")
            reset_auto_increment("evaluacionescl_registroadmin")
            reset_auto_increment("evaluacionescl_evaluacionlectura")
            reset_auto_increment("evaluacionescl_evaluacionlecturaindividual")
            reset_auto_increment("evaluacionescl_vistaadmin")

            # Guardar mensaje antes de limpiar sesión
            add_message(request, messages.SUCCESS, "⚠️ Se reseteó TODO, incluyendo el admin en sesión.")

            # Limpiar la sesión y redirigir
            request.session.flush()
            return redirect("login_admin")

        elif "cancelar" in request.POST:
            messages.info(request, "🚫 Operación cancelada.")
            return redirect("dashboard_admin")

    return render(request, "evaluacionescl/resetear_datos_confirmacion.html")